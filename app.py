import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

pd.set_option("styler.render.max_elements", 2**31 - 1)

# ── Constants ────────────────────────────────────────────────────────────────
MONTH_MAP = {
    "Січ": 1, "Лют": 2, "Бер": 3, "Кві": 4,
    "Тра": 5, "Чер": 6, "Лип": 7, "Сер": 8,
    "Вер": 9, "Жов": 10, "Лис": 11, "Гру": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
MONTH_LABELS = {
    1: "jan", 2: "feb", 3: "mar", 4: "apr",
    5: "may", 6: "jun", 7: "jul", 8: "aug",
    9: "sep", 10: "oct", 11: "nov", 12: "dec",
}
MONTHS_LIST = [MONTH_LABELS[m] for m in range(1, 13)]

PURPLE    = "#5b2d8e"
GREY      = "#c0c0c0"
RED_LINE  = "#c0392b"
YELLOW    = "#f0c000"
GREEN_HDR = "#2e7d32"
TEAL      = "#0d7377"
TEAL_HDR  = "#085f63"
ORANGE    = "#e67e22"


# ── Helpers ───────────────────────────────────────────────────────────────────
def get_month_num(series):
    if pd.api.types.is_numeric_dtype(series):
        return series.astype(int)
    return series.astype(str).str.strip().map(MONTH_MAP).fillna(0).astype(int)


@st.cache_data
def load_excel(file_bytes, file_name, sheet_name):
    import os
    buf = io.BytesIO(file_bytes)
    ext = os.path.splitext(file_name)[1].lower()
    engine = "pyxlsb" if ext == ".xlsb" else None
    df = pd.read_excel(buf, sheet_name=sheet_name, **({} if engine is None else {"engine": engine}))
    df.columns = df.columns.str.strip()
    return df


def _prep(df, col_month):
    """Add _m column in-place."""
    out = df.copy()
    out["_m"] = get_month_num(out[col_month])
    return out


def _fact_rows(df, col_plf):
    return df[df[col_plf] == "F"] if col_plf and col_plf in df.columns else df


def _plan_rows(df, col_plf):
    if col_plf and col_plf in df.columns:
        return df[df[col_plf] == "PL"]
    return pd.DataFrame(columns=df.columns)


# ── Core calculation: build global average respecting group_factors ───────────
#
# LOGIC (identical for both absolute and ratio columns):
#   1. From the FULL dataset (all TTs), compute per-TT average over
#      (group_factors + article) grouping — same as absolute Average.
#   2. Sum (absolute) or mean (ratio) those per-TT values for the selected TTs
#      to get the "dynamic average" series across months.
#
# This mirrors exactly how build_article_monthly computes Average.
# ─────────────────────────────────────────────────────────────────────────────

def _build_global_avg(df_all_fact, col_value, col_article, col_tt,
                      group_factors, agg_fn="mean"):
    """
    Compute global average per (group_factors + article) from the full fact dataset.
    agg_fn: 'mean' for ratio, 'mean' for absolute (both use mean at TT level).
    Returns DataFrame with Average_Calc column.
    """
    if group_factors:
        grp = list(dict.fromkeys(group_factors + [col_article]))
    else:
        grp = [col_tt, col_article, "_m"]

    return (
        df_all_fact
        .groupby(grp, as_index=False)[col_value]
        .agg(Average_Calc=agg_fn)
    )


def _merge_avg(tt_table, global_avg, group_factors, col_article, col_tt):
    """Merge global average onto tt_table."""
    if group_factors:
        on = list(dict.fromkeys(group_factors + [col_article]))
    else:
        on = [col_tt, col_article, "_m"]
    return pd.merge(tt_table, global_avg, on=on, how="left")


def build_article_monthly(df, df_filtered, col_tt, col_article, col_month,
                           col_value, col_plf, selected_art, selected_tts, group_factors):
    """Absolute value monthly table: Plan / Fact / Average / Delta."""
    art_all  = _prep(df[df[col_article] == selected_art], col_month)
    art_filt = _prep(df_filtered[df_filtered[col_article] == selected_art], col_month)

    if art_filt.empty:
        return pd.DataFrame(0.0, index=range(1, 13), columns=["Plan", "Fact", "Average", "Delta"])

    if not selected_tts:
        selected_tts = art_filt[col_tt].dropna().unique().tolist()

    plan = (_plan_rows(art_filt, col_plf)
            .groupby("_m")[col_value].sum()
            .reindex(range(1, 13), fill_value=0).rename("Plan"))
    fact = (_fact_rows(art_filt, col_plf)
            .groupby("_m")[col_value].sum()
            .reindex(range(1, 13), fill_value=0).rename("Fact"))

    # Global average from full dataset
    all_fact = _fact_rows(art_all, col_plf)
    global_avg = _build_global_avg(all_fact, col_value, col_article, col_tt, group_factors)

    tt_grp = list(dict.fromkeys([col_tt] + group_factors + ["_m", col_article]))
    tt_table = (
        _fact_rows(art_filt, col_plf)
        .groupby(tt_grp, as_index=False)[col_value]
        .sum()
        .rename(columns={col_value: "Fact"})
    )
    tt_table = _merge_avg(tt_table, global_avg, group_factors, col_article, col_tt)
    tt_table["Fact"]         = tt_table["Fact"].fillna(0)
    tt_table["Average_Calc"] = tt_table["Average_Calc"].fillna(0)
    tt_table.loc[tt_table["Fact"] == 0, "Average_Calc"] = 0

    dynamic_average = (
        tt_table[tt_table[col_tt].isin(selected_tts)]
        .groupby("_m")["Average_Calc"].sum()
        .reindex(range(1, 13), fill_value=0)
        .rename("Average")
    )

    merged = pd.DataFrame(index=range(1, 13)).join(plan).join(fact).join(dynamic_average).fillna(0)
    merged.index.name = "month"
    merged.loc[merged["Fact"] == 0, "Average"] = 0
    merged["Delta"] = merged["Fact"] - merged["Average"]
    return merged


def build_ratio_monthly(df_filtered, col_tt, col_article, col_month,
                         col_ratio, col_plf, selected_art, selected_tts,
                         df_all=None, group_factors=None):
    """
    Ratio (% в ТО) monthly table: Plan / Fact / Average / Delta.
    Average is computed EXACTLY like absolute Average:
      - group by (group_factors + article) on full dataset → mean per TT
      - sum selected TTs per month
    """
    if group_factors is None:
        group_factors = []

    src_all  = (df_all if df_all is not None else df_filtered)
    art_all  = _prep(src_all[src_all[col_article] == selected_art].copy(), col_month)
    art_filt = _prep(df_filtered[df_filtered[col_article] == selected_art].copy(), col_month)

    for d in (art_all, art_filt):
        d[col_ratio] = pd.to_numeric(d[col_ratio], errors="coerce")

    if art_filt.empty:
        return pd.DataFrame(0.0, index=range(1, 13), columns=["Plan", "Fact", "Average", "Delta"])

    if not selected_tts:
        selected_tts = art_filt[col_tt].dropna().unique().tolist()

    fact_src = _fact_rows(art_filt, col_plf)
    plan_src = _plan_rows(art_filt, col_plf)

    fact = (fact_src.groupby("_m")[col_ratio].mean()
            .reindex(range(1, 13), fill_value=np.nan).rename("Fact"))
    plan = (plan_src.groupby("_m")[col_ratio].mean()
            .reindex(range(1, 13), fill_value=np.nan).rename("Plan"))

    # ── Average: same logic as absolute ──────────────────────────────────────
    all_fact = _fact_rows(art_all, col_plf)
    global_avg = _build_global_avg(all_fact, col_ratio, col_article, col_tt,
                                    group_factors, agg_fn="mean")

    tt_grp = list(dict.fromkeys([col_tt] + group_factors + ["_m", col_article]))
    tt_table = (
        fact_src.groupby(tt_grp, as_index=False)[col_ratio]
        .mean()
        .rename(columns={col_ratio: "Fact_tt"})
    )
    tt_table = _merge_avg(tt_table, global_avg, group_factors, col_article, col_tt)
    tt_table["Fact_tt"]      = tt_table["Fact_tt"].fillna(0)
    tt_table["Average_Calc"] = tt_table["Average_Calc"].fillna(0)
    tt_table.loc[tt_table["Fact_tt"] == 0, "Average_Calc"] = 0

    # For ratio: use mean (not sum) when aggregating selected TTs
    dynamic_average = (
        tt_table[tt_table[col_tt].isin(selected_tts)]
        .groupby("_m")["Average_Calc"].mean()
        .reindex(range(1, 13), fill_value=0)
        .rename("Average")
    )

    merged = (pd.DataFrame(index=range(1, 13))
              .join(plan).join(fact).join(dynamic_average)
              .fillna(0.0))
    merged.index.name = "month"
    merged.loc[merged["Fact"] == 0, "Average"] = 0
    merged["Delta"] = merged["Fact"] - merged["Average"]
    return merged


# ── Heatmap builders ─────────────────────────────────────────────────────────

def build_heat_data(df, df_filtered, col_tt, col_article, col_month, col_value,
                    col_plf, group_factors, articles_to_show, mode):
    df_num = df.copy()
    df_num[col_value] = pd.to_numeric(df_num[col_value], errors="coerce")

    global_avg_std = (
        _fact_rows(df_num, col_plf)
        .groupby(group_factors + [col_article], as_index=False)[col_value]
        .agg(Average_Calc="mean", Std="std")
    )
    filt = df_filtered.copy()
    filt[col_value] = pd.to_numeric(filt[col_value], errors="coerce")
    data_heat = _prep(
        filt[(filt[col_plf] == "F") & (filt[col_article].isin(articles_to_show))],
        col_month
    )

    tt_grp = list(dict.fromkeys([col_tt] + group_factors + ["_m", col_article]))
    tt_table = (
        data_heat.groupby(tt_grp, as_index=False)[col_value]
        .sum().rename(columns={col_value: "Fact"})
    )
    merge_cols = list(dict.fromkeys(group_factors + [col_article]))
    tt_table = pd.merge(tt_table, global_avg_std, on=merge_cols, how="left")
    tt_table["Delta"]   = tt_table["Fact"] - tt_table["Average_Calc"]
    tt_table["Delta_%"] = tt_table["Delta"] / tt_table["Average_Calc"].replace(0, np.nan)
    tt_table["Z"]       = tt_table["Delta"] / tt_table["Std"].replace(0, np.nan)

    val_col = {"Delta": "Delta", "Delta %": "Delta_%", "Z-score": "Z",
               "Fact": "Fact", "Average": "Average_Calc"}[mode]

    heat = tt_table.pivot_table(index=col_tt, columns="_m", values=val_col, aggfunc="sum")
    heat = _fill_heat_cols(heat)
    heat["РАЗОМ"] = heat.sum(axis=1, numeric_only=True)
    return heat, tt_table, val_col


def build_ratio_heat_data(df, df_filtered, col_tt, col_article, col_month,
                           col_ratio, col_plf, articles_to_show, mode,
                           group_factors=None):
    if group_factors is None:
        group_factors = []

    df_num = df_filtered.copy()
    df_num[col_ratio] = pd.to_numeric(df_num[col_ratio], errors="coerce")
    df_num = _prep(df_num, col_month)

    has_plf = col_plf and col_plf in df_num.columns
    data_heat = df_num[
        (df_num[col_plf] == "F" if has_plf else True) &
        df_num[col_article].isin(articles_to_show)
    ].copy() if has_plf else df_num[df_num[col_article].isin(articles_to_show)].copy()

    # Global average — same logic as build_ratio_monthly
    df_all_num = df.copy()
    df_all_num[col_ratio] = pd.to_numeric(df_all_num[col_ratio], errors="coerce")
    df_all_num = _prep(df_all_num, col_month)
    avg_src = (_fact_rows(df_all_num, col_plf) if has_plf else df_all_num)
    avg_src = avg_src[avg_src[col_article].isin(articles_to_show)]

    if group_factors:
        grp_cols = list(dict.fromkeys(group_factors + [col_article, "_m"]))
        global_avg = (avg_src.groupby(grp_cols)[col_ratio]
                      .mean().reset_index()
                      .rename(columns={col_ratio: "Average_Calc"}))
        merge_on = grp_cols
    else:
        global_avg = (avg_src.groupby([col_article, "_m"])[col_ratio]
                      .mean().reset_index()
                      .rename(columns={col_ratio: "Average_Calc"}))
        merge_on = [col_article, "_m"]

    tt_grp = list(dict.fromkeys([col_tt] + group_factors + [col_article, "_m"]))
    tt_table = (
        data_heat.groupby(tt_grp, as_index=False)[col_ratio]
        .mean().rename(columns={col_ratio: "Fact"})
    )
    tt_table = pd.merge(tt_table, global_avg, on=merge_on, how="left")
    tt_table["Delta"]   = tt_table["Fact"] - tt_table["Average_Calc"]
    tt_table["Delta_%"] = tt_table["Delta"] / tt_table["Average_Calc"].replace(0, np.nan)
    tt_table["Std"]     = np.nan
    tt_table["Z"]       = np.nan

    val_col = {"Delta": "Delta", "Delta %": "Delta_%",
               "Fact": "Fact", "Average": "Average_Calc"}.get(mode, "Delta")

    heat = tt_table.pivot_table(index=col_tt, columns="_m", values=val_col, aggfunc="mean")
    heat = _fill_heat_cols(heat)
    heat["РАЗОМ"] = heat.mean(axis=1, numeric_only=True)
    return heat, tt_table, val_col


def _fill_heat_cols(heat):
    for m in range(1, 13):
        if m not in heat.columns:
            heat[m] = None
    heat = heat[sorted(heat.columns)]
    heat.columns = [MONTH_LABELS.get(int(c), str(c)) for c in heat.columns]
    return heat


# ── TT Pivot ─────────────────────────────────────────────────────────────────

def build_tt_pivot(df_filtered, col_tt, col_article, col_month, col_value,
                   col_plf, articles_to_show):
    df_num = _prep(df_filtered.copy(), col_month)
    df_num[col_value] = pd.to_numeric(df_num[col_value], errors="coerce")

    all_tts = sorted(df_filtered[col_tt].dropna().unique(), key=str)
    rows = []
    for tt in all_tts:
        sub = df_num[df_num[col_tt] == tt]
        for article in articles_to_show:
            sub_a  = sub[sub[col_article] == article]
            plan_m = _plan_rows(sub_a, col_plf).groupby("_m")[col_value].sum()
            fact_m = _fact_rows(sub_a, col_plf).groupby("_m")[col_value].sum()
            row    = {"ТТ": tt, "Стаття": article}
            for m in range(1, 13):
                ml = MONTH_LABELS[m]
                row[f"plan_{ml}"] = plan_m.get(m, 0)
                row[f"fact_{ml}"] = fact_m.get(m, 0)
            row["Plan_РАЗОМ"]  = sum(plan_m.get(m, 0) for m in range(1, 13))
            row["Fact_РАЗОМ"]  = sum(fact_m.get(m, 0) for m in range(1, 13))
            row["Delta_РАЗОМ"] = row["Fact_РАЗОМ"] - row["Plan_РАЗОМ"]
            row["Pct_РАЗОМ"]   = (
                (row["Fact_РАЗОМ"] / row["Plan_РАЗОМ"] - 1) * 100
                if row["Plan_РАЗОМ"] != 0 else None
            )
            rows.append(row)

    if not rows:
        return pd.DataFrame()

    df_tt = pd.DataFrame(rows)
    num_cols = ([f"plan_{MONTH_LABELS[m]}" for m in range(1, 13)] +
                [f"fact_{MONTH_LABELS[m]}" for m in range(1, 13)] +
                ["Plan_РАЗОМ", "Fact_РАЗОМ", "Delta_РАЗОМ"])

    df_agg = df_tt.groupby("ТТ")[num_cols].sum().reset_index()
    df_agg["Pct_РАЗОМ"] = df_agg.apply(
        lambda r: (r["Fact_РАЗОМ"] / r["Plan_РАЗОМ"] - 1) * 100
        if r["Plan_РАЗОМ"] != 0 else None, axis=1
    )
    for m in range(1, 13):
        ml = MONTH_LABELS[m]
        df_agg[f"delta_{ml}"] = df_agg[f"fact_{ml}"] - df_agg[f"plan_{ml}"]
        df_agg[f"pct_{ml}"] = df_agg.apply(
            lambda r, ml=ml: (r[f"fact_{ml}"] / r[f"plan_{ml}"] - 1) * 100
            if r[f"plan_{ml}"] != 0 else None, axis=1
        )
    return df_agg


# ── HTML / Styling helpers ────────────────────────────────────────────────────

def _th(bg_color):
    return (f"background:{bg_color};color:white;font-weight:bold;border:1px solid #aaa;"
            "padding:4px 8px;text-align:center;font-size:0.78rem;")


TD  = "border:1px solid #ccc;padding:3px 7px;text-align:right;font-size:0.78rem;"
TL  = "border:1px solid #ccc;padding:3px 7px;font-size:0.78rem;font-weight:600;white-space:nowrap;"


def _make_pills(series, color, bg):
    pills = ""
    for tt, val in series.items():
        sign    = "+" if val > 0 else ""
        val_fmt = f"{val:,.0f}".replace(",", " ")
        pills  += (
            f'<span style="display:inline-block;background:{bg};color:{color};'
            f'border-radius:4px;padding:2px 9px;margin:2px 3px;font-size:0.75rem;'
            f'font-weight:600;white-space:nowrap;">'
            f'{tt}&nbsp;<span style="opacity:.7;font-weight:400;">({sign}{val_fmt})</span></span>'
        )
    return pills


def _make_pct_pills(series, color, bg):
    pills = ""
    for tt, val in series.items():
        sign = "+" if val > 0 else ""
        pills += (
            f'<span style="display:inline-block;background:{bg};color:{color};'
            f'border-radius:4px;padding:2px 9px;margin:2px 3px;font-size:0.75rem;'
            f'font-weight:600;white-space:nowrap;">'
            f'{tt}&nbsp;<span style="opacity:.7;font-weight:400;">({sign}{val:.2f}%)</span></span>'
        )
    return pills


def _render_slicer(article_idx, prefix, df_filtered, col_tt, col_article, title):
    """Shared TT slicer widget. Returns current active_tt."""
    skey = f"{prefix}_slicer_tt_{article_idx}"
    if skey not in st.session_state:
        st.session_state[skey] = "__ALL__"
    active_tt = st.session_state[skey]

    available_tts = sorted(
        df_filtered[df_filtered[col_article] == title][col_tt].dropna().unique(), key=str
    )
    if not available_tts:
        return active_tt

    with st.expander("🏪 Слайсер по ТТ — клікни для деталізації", expanded=False):
        search_key = f"{prefix}_slicer_search_{article_idx}"
        search_val = st.text_input("🔎 Пошук магазину", value="",
                                   placeholder="Введіть назву...", key=search_key)
        filtered_tts = (
            [t for t in available_tts if search_val.lower() in str(t).lower()]
            if search_val else available_tts
        )
        all_options   = ["__ALL__"] + list(filtered_tts)
        VISIBLE_ITEMS = 4 * 6
        show_all_key  = f"{prefix}_slicer_showall_{article_idx}"
        if show_all_key not in st.session_state:
            st.session_state[show_all_key] = False
        items_to_show = all_options if st.session_state[show_all_key] else all_options[:VISIBLE_ITEMS]

        for row_start in range(0, len(items_to_show), 6):
            chunk = items_to_show[row_start:row_start + 6]
            cols  = st.columns(len(chunk))
            for ci, tt_opt in enumerate(chunk):
                label    = "🔁 Всі" if tt_opt == "__ALL__" else str(tt_opt)
                btn_type = "primary" if active_tt == tt_opt else "secondary"
                with cols[ci]:
                    if st.button(label, key=f"{prefix}_btn_{article_idx}_{row_start}_{ci}_{hash(str(tt_opt))}",
                                 type=btn_type, use_container_width=True):
                        st.session_state[skey] = tt_opt
                        st.rerun()

        if len(all_options) > VISIBLE_ITEMS:
            remaining = len(all_options) - VISIBLE_ITEMS
            label = "▲ Згорнути" if st.session_state[show_all_key] else f"▼ Показати ще {remaining}"
            if st.button(label, key=f"{prefix}_toggle_{article_idx}", use_container_width=False):
                st.session_state[show_all_key] = not st.session_state[show_all_key]
                st.rerun()

        if active_tt != "__ALL__":
            st.caption(f"📍 Показано тільки: **{active_tt}**")
        else:
            st.caption(f"Показано всі ТТ · знайдено: {len(filtered_tts)}")

    return active_tt


# ── Article block — absolute ─────────────────────────────────────────────────

def render_article_block(title, table_df, df, df_filtered,
                          col_tt, col_article, col_month, col_value, col_plf,
                          group_factors, tt_val, article_idx):
    rows_cfg = [
        ("План",    "Plan",    "#ffffff", "#333333"),
        ("Факт",    "Fact",    "#e8d5f5", PURPLE),
        ("Average", "Average", "#fde8e8", RED_LINE),
        ("Дельта",  "Delta",   "#fff9e0", "#b8860b"),
    ]
    th = _th(GREEN_HDR)

    active_tt = _render_slicer(article_idx, "abs", df_filtered, col_tt, col_article, title)

    if active_tt != "__ALL__":
        df_filt_tt = df_filtered[df_filtered[col_tt] == active_tt].copy()
        display_df = build_article_monthly(
            df, df_filt_tt, col_tt, col_article, col_month, col_value,
            col_plf, title, [active_tt], group_factors
        )
    else:
        display_df = table_df

    badge = (f'<span style="margin-left:10px;background:{PURPLE};color:white;font-size:0.78rem;'
             f'padding:2px 10px;border-radius:10px;">📍 {active_tt}</span>'
             if active_tt != "__ALL__" else "")
    st.markdown(f"""
    <div style="margin-top:20px;margin-bottom:4px;">
      <span style="background:{GREEN_HDR};color:white;font-weight:700;padding:4px 14px;
                   font-size:0.9rem;border-radius:2px;">{title}</span>{badge}
    </div>""", unsafe_allow_html=True)

    html = (f'<div style="overflow-x:auto;"><table style="border-collapse:collapse;width:100%;'
            f'margin-bottom:6px;"><thead><tr>'
            f'<th style="{th}">Показник</th>'
            + "".join(f'<th style="{th}">{m}</th>' for m in MONTHS_LIST)
            + f'<th style="{th}">Разом</th></tr></thead><tbody>')
    for label, col, bg, color in rows_cfg:
        vals  = [display_df.loc[m, col] for m in range(1, 13)]
        total = sum(vals)
        html += f'<tr style="background:{bg};"><td style="{TL}color:{color};">{label}</td>'
        for v in vals:
            html += f'<td style="{TD}{"color:#c0392b;" if v < 0 else ""}">{v:,.0f}</td>'
        html += f'<td style="{TD}font-weight:700;">{total:,.0f}</td></tr>'
    html += "</tbody></table></div>"
    st.markdown(html, unsafe_allow_html=True)

    # Metrics
    facts       = [display_df.loc[m, "Fact"] for m in range(1, 13)]
    nz          = [f for f in facts if f != 0]
    avg_monthly = np.mean(nz) if nz else 0
    total_fact  = sum(facts)
    total_plan  = sum(display_df.loc[m, "Plan"] for m in range(1, 13))
    total_delta = total_fact - total_plan
    pct         = ((total_fact / total_plan - 1) * 100) if total_plan != 0 else None
    pct_str     = (f"{'+' if pct >= 0 else ''}{pct:.1f}%" if pct is not None else "—")
    pct_color   = RED_LINE if (pct or 0) > 0 else GREEN_HDR
    delta_color = RED_LINE if total_delta > 0 else GREEN_HDR

    best_pills = worst_pills = ""
    if active_tt == "__ALL__":
        sub = df_filtered[(df_filtered[col_article] == title) & (df_filtered[col_plf] == "F")].copy()
        sub[col_value] = pd.to_numeric(sub[col_value], errors="coerce")
        if not sub.empty and col_tt in sub.columns:
            tt_totals = sub.groupby(col_tt)[col_value].sum().dropna().sort_values()
            n = min(3, len(tt_totals))
            best_pills  = _make_pills(tt_totals.head(n), "#1b5e20", "#e8f5e9")
            worst_pills = _make_pills(tt_totals.tail(n).iloc[::-1], "#7f0000", "#ffebee")

    best_block = f"""
      <div style="flex:1;min-width:220px;">
        <div style="color:#888;font-size:0.71rem;margin-bottom:4px;text-transform:uppercase;">✅ Найекономніші магазини (мін. Fact)</div>
        <div>{best_pills or '<span style="color:#aaa;font-size:0.75rem;">немає даних</span>'}</div>
      </div>
      <div style="flex:1;min-width:220px;">
        <div style="color:#888;font-size:0.71rem;margin-bottom:4px;text-transform:uppercase;">❌ Найбільш витратні магазини (макс. Fact)</div>
        <div>{worst_pills or '<span style="color:#aaa;font-size:0.75rem;">немає даних</span>'}</div>
      </div>""" if active_tt == "__ALL__" else ""

    st.markdown(f"""
    <div style="display:flex;flex-wrap:wrap;gap:10px;background:#f9f6ff;
                border:1px solid #d0baf5;border-radius:6px;padding:10px 16px;margin:6px 0 10px 0;">
      <div style="min-width:130px;">
        <div style="color:#888;font-size:0.71rem;text-transform:uppercase;">Серед. Fact / міс.</div>
        <div style="font-size:1.1rem;font-weight:700;color:{PURPLE};">{avg_monthly:,.0f}</div>
      </div>
      <div style="min-width:130px;">
        <div style="color:#888;font-size:0.71rem;text-transform:uppercase;">Δ Fact − Plan</div>
        <div style="font-size:1.1rem;font-weight:700;color:{delta_color};">{('+' if total_delta > 0 else '')}{total_delta:,.0f}</div>
      </div>
      <div style="min-width:100px;">
        <div style="color:#888;font-size:0.71rem;text-transform:uppercase;">% до плану</div>
        <div style="font-size:1.1rem;font-weight:700;color:{pct_color};">{pct_str}</div>
      </div>
      {best_block}
    </div>""", unsafe_allow_html=True)

    fig = go.Figure([
        go.Bar(x=MONTHS_LIST, y=display_df["Plan"],    name="План",    marker_color=GREY),
        go.Bar(x=MONTHS_LIST, y=display_df["Fact"],    name="Факт",    marker_color=PURPLE),
        go.Scatter(x=MONTHS_LIST, y=display_df["Average"], name="Average",
                   line=dict(color=RED_LINE, width=3)),
        go.Scatter(x=MONTHS_LIST, y=display_df["Delta"],   name="Дельта",
                   line=dict(color=YELLOW, dash="dot")),
    ])
    fig.update_layout(height=320, margin=dict(t=30, b=50, l=10, r=10),
                      barmode="group", hovermode="x unified",
                      legend=dict(orientation="h", y=-0.18, x=0))
    st.plotly_chart(fig, use_container_width=True, key=f"chart_{article_idx}_{active_tt}")


# ── Article block — ratio ────────────────────────────────────────────────────

def render_ratio_article_block(title, table_df, df, df_filtered,
                                col_tt, col_article, col_month, col_ratio, col_plf,
                                tt_val, article_idx, group_factors=None):
    if group_factors is None:
        group_factors = []

    rows_cfg = [
        ("План %",    "Plan",    "#e8f8f8", TEAL_HDR),
        ("Факт %",    "Fact",    "#d0f0f0", TEAL),
        ("Average %", "Average", "#fde8e8", RED_LINE),
        ("Дельта %",  "Delta",   "#fff9e0", ORANGE),
    ]
    th = _th(TEAL_HDR)

    active_tt = _render_slicer(article_idx, "rat", df_filtered, col_tt, col_article, title)

    if active_tt != "__ALL__":
        df_filt_tt = df_filtered[df_filtered[col_tt] == active_tt].copy()
        display_df = build_ratio_monthly(
            df_filt_tt, col_tt, col_article, col_month, col_ratio, col_plf,
            title, [active_tt], df_all=df, group_factors=group_factors
        )
    else:
        display_df = table_df

    badge = (f'<span style="margin-left:10px;background:{TEAL};color:white;font-size:0.78rem;'
             f'padding:2px 10px;border-radius:10px;">📍 {active_tt}</span>'
             if active_tt != "__ALL__" else "")
    st.markdown(f"""
    <div style="margin-top:12px;margin-bottom:4px;">
      <span style="background:{TEAL_HDR};color:white;font-weight:700;padding:4px 14px;
                   font-size:0.85rem;border-radius:2px;">📊 % в ТО — {title}</span>{badge}
    </div>""", unsafe_allow_html=True)

    html = (f'<div style="overflow-x:auto;"><table style="border-collapse:collapse;width:100%;'
            f'margin-bottom:6px;"><thead><tr>'
            f'<th style="{th}">Показник</th>'
            + "".join(f'<th style="{th}">{m}</th>' for m in MONTHS_LIST)
            + f'<th style="{th}">Серед.</th></tr></thead><tbody>')
    for label, col, bg, color in rows_cfg:
        vals    = [display_df.loc[m, col] for m in range(1, 13)]
        nz      = [v for v in vals if v != 0]
        summary = np.mean(nz) if nz else 0.0
        html   += f'<tr style="background:{bg};"><td style="{TL}color:{color};">{label}</td>'
        for v in vals:
            html += f'<td style="{TD}{"color:#c0392b;" if v < 0 else ""}">{v:.2f}%</td>'
        html += f'<td style="{TD}font-weight:700;">{summary:.2f}%</td></tr>'
    html += "</tbody></table></div>"
    st.markdown(html, unsafe_allow_html=True)

    facts       = [display_df.loc[m, "Fact"] for m in range(1, 13)]
    nz          = [f for f in facts if f != 0]
    avg_monthly = np.mean(nz) if nz else 0.0
    avg_global  = np.mean([display_df.loc[m, "Average"] for m in range(1, 13)
                            if display_df.loc[m, "Average"] != 0] or [0])
    total_delta = np.mean([display_df.loc[m, "Delta"] for m in range(1, 13)
                            if display_df.loc[m, "Fact"] != 0] or [0])
    delta_color = RED_LINE if total_delta > 0 else GREEN_HDR

    best_pills = worst_pills = ""
    if active_tt == "__ALL__":
        sub = df_filtered[df_filtered[col_article] == title].copy()
        sub[col_ratio] = pd.to_numeric(sub[col_ratio], errors="coerce")
        if col_plf and col_plf in sub.columns:
            sub = sub[sub[col_plf] == "F"]
        if not sub.empty and col_tt in sub.columns:
            tt_avgs = sub.groupby(col_tt)[col_ratio].mean().dropna().sort_values()
            n = min(3, len(tt_avgs))
            best_pills  = _make_pct_pills(tt_avgs.head(n), "#1b5e20", "#e8f5e9")
            worst_pills = _make_pct_pills(tt_avgs.tail(n).iloc[::-1], "#7f0000", "#ffebee")

    best_block = f"""
      <div style="flex:1;min-width:220px;">
        <div style="color:#555;font-size:0.71rem;text-transform:uppercase;">✅ Найекономніші магазини (мін. %)</div>
        <div>{best_pills or '<span style="color:#aaa;font-size:0.75rem;">немає даних</span>'}</div>
      </div>
      <div style="flex:1;min-width:220px;">
        <div style="color:#555;font-size:0.71rem;text-transform:uppercase;">❌ Найбільш витратні магазини (макс. %)</div>
        <div>{worst_pills or '<span style="color:#aaa;font-size:0.75rem;">немає даних</span>'}</div>
      </div>""" if active_tt == "__ALL__" else ""

    st.markdown(f"""
    <div style="display:flex;flex-wrap:wrap;gap:10px;background:#e8f8f8;
                border:1px solid #8ecece;border-radius:6px;padding:10px 16px;margin:6px 0 10px 0;">
      <div style="min-width:130px;">
        <div style="color:#555;font-size:0.71rem;text-transform:uppercase;">Серед. Fact % / міс.</div>
        <div style="font-size:1.1rem;font-weight:700;color:{TEAL};">{avg_monthly:.2f}%</div>
      </div>
      <div style="min-width:130px;">
        <div style="color:#555;font-size:0.71rem;text-transform:uppercase;">Average % (норма)</div>
        <div style="font-size:1.1rem;font-weight:700;color:{RED_LINE};">{avg_global:.2f}%</div>
      </div>
      <div style="min-width:130px;">
        <div style="color:#555;font-size:0.71rem;text-transform:uppercase;">Δ Fact − Average</div>
        <div style="font-size:1.1rem;font-weight:700;color:{delta_color};">{('+' if total_delta > 0 else '')}{total_delta:.2f}%</div>
      </div>
      {best_block}
    </div>""", unsafe_allow_html=True)

    fig = go.Figure([
        go.Bar(x=MONTHS_LIST, y=display_df["Plan"],    name="План %",  marker_color="#8ecece", opacity=0.8),
        go.Bar(x=MONTHS_LIST, y=display_df["Fact"],    name="Факт %",  marker_color=TEAL, opacity=0.95),
        go.Scatter(x=MONTHS_LIST, y=display_df["Average"], name="Average %",
                   line=dict(color=RED_LINE, width=3)),
        go.Scatter(x=MONTHS_LIST, y=display_df["Delta"],   name="Δ %",
                   line=dict(color=ORANGE, dash="dot"), yaxis="y2"),
    ])
    fig.update_layout(
        height=320, margin=dict(t=30, b=50, l=10, r=10),
        barmode="group", hovermode="x unified",
        yaxis=dict(tickformat=".2f", ticksuffix="%"),
        yaxis2=dict(overlaying="y", side="right", tickformat=".2f",
                    ticksuffix="%", showgrid=False),
        legend=dict(orientation="h", y=-0.18, x=0),
    )
    st.plotly_chart(fig, use_container_width=True, key=f"ratio_chart_{article_idx}_{active_tt}")


# ── Ratio heatmap section ─────────────────────────────────────────────────────

def render_ratio_heatmap_section(df, df_filtered, col_tt, col_article,
                                  col_month, col_ratio, col_plf,
                                  articles_to_show, ratio_mode, group_factors=None):
    if group_factors is None:
        group_factors = []

    heat, tt_table, val_col = build_ratio_heat_data(
        df, df_filtered, col_tt, col_article, col_month,
        col_ratio, col_plf, articles_to_show, ratio_mode,
        group_factors=group_factors,
    )

    st.markdown(f"""
    <div style="margin-bottom:6px;">
      <span style="background:{TEAL_HDR};color:white;font-weight:700;
                   padding:4px 14px;font-size:0.9rem;border-radius:2px;">
        🌡️ Карта аномалій % в ТО
      </span>
      <span style="margin-left:12px;color:#666;font-size:0.8rem;">
        Режим: <b>{ratio_mode}</b> · значення у відсотках
      </span>
    </div>""", unsafe_allow_html=True)

    st.dataframe(
        heat.style
            .background_gradient(cmap="RdYlGn_r", axis=None)
            .highlight_null(color="white")
            .format(lambda v: f"{v:.2f}%" if pd.notna(v) else "", na_rep=""),
        use_container_width=True,
    )

    st.markdown(f"""
    <div style="margin:14px 0 6px 0;">
      <span style="background:{TEAL_HDR};color:white;font-weight:700;
                   padding:4px 14px;font-size:0.9rem;border-radius:2px;">
        🏆 TOP / ANTITOP магазинів — % в ТО
      </span>
    </div>""", unsafe_allow_html=True)

    sum_val = tt_table.groupby(col_tt)[val_col].mean().reset_index()
    n_tt    = st.slider("Кількість магазинів (% в ТО)", 1, 100, 10, key="ratio_n_tt_slider")
    top     = sum_val.sort_values(val_col, ascending=True).head(n_tt)
    antitop = sum_val.sort_values(val_col, ascending=False).head(n_tt)
    fmt_fn  = lambda v: f"{v:.2f}%" if pd.notna(v) else "-"

    ca, cb = st.columns(2)
    with ca:
        st.write("✅ Top (найменший %)")
        st.dataframe(top.style.background_gradient(cmap="RdYlGn", subset=[val_col])
                        .format({val_col: fmt_fn}), use_container_width=True)
    with cb:
        st.write("❌ Antitop (найбільший %)")
        st.dataframe(antitop.style.background_gradient(cmap="RdYlGn_r", subset=[val_col])
                             .format({val_col: fmt_fn}), use_container_width=True)


# ── Excel Export ──────────────────────────────────────────────────────────────

def export_excel(df, df_filtered, col_tt, col_article, col_month, col_value,
                 col_plf, articles_to_show, tt_val, group_factors, metric_col,
                 mode, pivot_df, df_tt_agg=None, col_ratio=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    NUM_FMT = '# ##0;-# ##0;-'
    PCT_FMT = '+0.0%;-0.0%;-'

    def hdr_fill(h):
        h = h.lstrip("#")
        return PatternFill("solid", start_color=h, end_color=h)

    def thin_border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def scw(ws, ci, w):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Зведена таблиця ──────────────────────────────────────────────────
    ws_p = wb.create_sheet("Зведена_таблиця")
    ws_p.freeze_panes = "B2"
    header = ["Стаття"] + MONTHS_LIST + ["РАЗОМ"]
    for ci, h in enumerate(header, 1):
        c = ws_p.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = hdr_fill("2e7d32")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws_p.row_dimensions[1].height = 28
    ws_p.column_dimensions["A"].width = 36
    for ci in range(2, len(header) + 1):
        scw(ws_p, ci, 11)

    for ri, article in enumerate(articles_to_show, 2):
        tdf  = build_article_monthly(df, df_filtered, col_tt, col_article,
                                     col_month, col_value, col_plf, article, tt_val, group_factors)
        vals = [article] + [tdf.loc[m, metric_col] for m in range(1, 13)]
        vals.append(sum(tdf.loc[m, metric_col] for m in range(1, 13)))
        for ci, v in enumerate(vals, 1):
            c = ws_p.cell(row=ri, column=ci, value=v)
            c.border = thin_border()
            c.font   = Font(name="Arial", size=9)
            if ci == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.number_format = NUM_FMT
                c.alignment     = Alignment(horizontal="right")
                if isinstance(v, (int, float)) and v < 0:
                    c.font = Font(name="Arial", size=9, color="C0392B")

    # ── 2. Листи по статтях ─────────────────────────────────────────────────
    row_labels = ["План", "Факт", "Average", "Дельта"]
    row_keys   = ["Plan", "Fact", "Average", "Delta"]
    row_fills  = ["FFFFFF", "e8d5f5", "fde8e8", "fff9e0"]
    row_colors = ["333333", "5b2d8e", "c0392b", "b8860b"]

    for article in articles_to_show:
        tdf  = build_article_monthly(df, df_filtered, col_tt, col_article,
                                     col_month, col_value, col_plf, article, tt_val, group_factors)
        safe = article[:28].replace("/", "_").replace("\\", "_")
        ws   = wb.create_sheet(safe)
        ws.freeze_panes = "B3"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        tc = ws.cell(row=1, column=1, value=article)
        tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        tc.fill = hdr_fill("5b2d8e")
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        headers = ["Показник"] + MONTHS_LIST + ["РАЗОМ"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill      = hdr_fill("2e7d32")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
        ws.row_dimensions[2].height = 18

        for ri, (label, key, fill_hex, color_hex) in enumerate(
                zip(row_labels, row_keys, row_fills, row_colors), 3):
            vals  = [tdf.loc[m, key] for m in range(1, 13)]
            total = sum(vals)
            nc = ws.cell(row=ri, column=1, value=label)
            nc.font = Font(bold=True, color=color_hex, name="Arial", size=9)
            nc.fill = hdr_fill(fill_hex)
            nc.border = thin_border()
            nc.alignment = Alignment(horizontal="left")
            for ci, v in enumerate(vals, 2):
                c = ws.cell(row=ri, column=ci, value=v)
                c.number_format = NUM_FMT
                c.fill   = hdr_fill(fill_hex)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="right")
                c.font = Font(name="Arial", size=9, color="C0392B" if v < 0 else color_hex)
            tc2 = ws.cell(row=ri, column=14, value=total)
            tc2.number_format = NUM_FMT
            tc2.fill   = hdr_fill(fill_hex)
            tc2.border = thin_border()
            tc2.alignment = Alignment(horizontal="right")
            tc2.font = Font(bold=True, name="Arial", size=9,
                            color="C0392B" if total < 0 else color_hex)

        ws.column_dimensions["A"].width = 12
        for ci in range(2, 15):
            scw(ws, ci, 11)

        fig = go.Figure([
            go.Bar(x=MONTHS_LIST, y=[tdf.loc[m, "Plan"] for m in range(1, 13)],
                   name="План", marker_color="#c0c0c0", opacity=0.9),
            go.Bar(x=MONTHS_LIST, y=[tdf.loc[m, "Fact"] for m in range(1, 13)],
                   name="Факт", marker_color="#5b2d8e", opacity=0.95),
            go.Scatter(x=MONTHS_LIST, y=[tdf.loc[m, "Average"] for m in range(1, 13)],
                       mode="lines+markers", name="Average",
                       line=dict(color="#c0392b", width=2.5), marker=dict(size=8)),
            go.Scatter(x=MONTHS_LIST, y=[tdf.loc[m, "Delta"] for m in range(1, 13)],
                       mode="lines+markers", name="Дельта",
                       line=dict(color="#f0c000", width=2), marker=dict(size=7), yaxis="y2"),
        ])
        fig.update_layout(
            barmode="group", height=320, width=900,
            plot_bgcolor="white", paper_bgcolor="white",
            title=dict(text=f"Аналіз — {article}", x=0.5, font=dict(size=12)),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#ececec"),
            yaxis2=dict(overlaying="y", side="right", showgrid=False),
            legend=dict(orientation="v", x=1.06, y=1, font=dict(size=9)),
            margin=dict(t=40, b=30, l=55, r=130), font=dict(family="Arial"),
        )
        try:
            import plotly.io as pio
            img_obj = XLImage(io.BytesIO(pio.to_image(fig, format="png", scale=1.5)))
            img_obj.anchor = "A7"
            ws.add_image(img_obj)
        except Exception:
            ws.cell(row=7, column=1, value="⚠️ Графік недоступний (pip install kaleido)")

    # ── 3. % в ТО ───────────────────────────────────────────────────────────
    if col_ratio:
        ws_ratio = wb.create_sheet("% в ТО_зведена")
        ws_ratio.freeze_panes = "B2"
        r_header = ["Стаття"] + MONTHS_LIST + ["Серед."]
        for ci, h in enumerate(r_header, 1):
            c = ws_ratio.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill = hdr_fill("085f63")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        ws_ratio.row_dimensions[1].height = 28
        ws_ratio.column_dimensions["A"].width = 36
        for ci in range(2, len(r_header) + 1):
            scw(ws_ratio, ci, 11)

        r_cfg = [
            ("План %",    "Plan",    "e8f8f8", "085f63"),
            ("Факт %",    "Fact",    "d0f0f0", "0d7377"),
            ("Average %", "Average", "fde8e8", "c0392b"),
            ("Δ %",       "Delta",   "fff9e0", "e67e22"),
        ]
        for art_ri, article in enumerate(articles_to_show):
            rdf = build_ratio_monthly(
                df_filtered, col_tt, col_article, col_month,
                col_ratio, col_plf, article, tt_val,
                df_all=df, group_factors=group_factors
            )
            base_row = 2 + art_ri * 5
            ws_ratio.merge_cells(start_row=base_row, start_column=1,
                                  end_row=base_row, end_column=len(r_header))
            tc = ws_ratio.cell(row=base_row, column=1, value=article)
            tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            tc.fill = hdr_fill("0d7377")
            tc.alignment = Alignment(horizontal="left", vertical="center")
            ws_ratio.row_dimensions[base_row].height = 18

            for sub_ri, (label, key, fill_hex, color_hex) in enumerate(r_cfg, base_row + 1):
                vals    = [rdf.loc[m, key] for m in range(1, 13)]
                nz      = [v for v in vals if v != 0]
                summary = np.mean(nz) if nz else 0.0
                nc = ws_ratio.cell(row=sub_ri, column=1, value=label)
                nc.font = Font(bold=True, color=color_hex, name="Arial", size=9)
                nc.fill = hdr_fill(fill_hex)
                nc.border = thin_border()
                for ci, v in enumerate(vals, 2):
                    c = ws_ratio.cell(row=sub_ri, column=ci, value=round(v, 4))
                    c.number_format = '0.00"%"'
                    c.fill = hdr_fill(fill_hex)
                    c.border = thin_border()
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(name="Arial", size=9,
                                  color="C0392B" if v < 0 else color_hex)
                sc = ws_ratio.cell(row=sub_ri, column=14, value=round(summary, 4))
                sc.number_format = '0.00"%"'
                sc.fill = hdr_fill(fill_hex)
                sc.border = thin_border()
                sc.alignment = Alignment(horizontal="right")
                sc.font = Font(bold=True, name="Arial", size=9,
                               color="C0392B" if summary < 0 else color_hex)

    # ── 4. ТТ-Зведена ───────────────────────────────────────────────────────
    if df_tt_agg is not None and not df_tt_agg.empty:
        ws_tt = wb.create_sheet("ТТ_Зведена")
        ws_tt.freeze_panes = "B2"

        tt_export_cols = (
            ["ТТ"]
            + [f"fact_{MONTH_LABELS[m]}" for m in range(1, 13)]
            + [f"plan_{MONTH_LABELS[m]}" for m in range(1, 13)]
            + [f"delta_{MONTH_LABELS[m]}" for m in range(1, 13)]
            + ["Fact_РАЗОМ", "Plan_РАЗОМ", "Delta_РАЗОМ", "Pct_РАЗОМ"]
        )
        tt_export_cols = [c for c in tt_export_cols if c in df_tt_agg.columns]

        tt_lbl = {"ТТ": "ТТ", "Fact_РАЗОМ": "Fact РАЗОМ", "Plan_РАЗОМ": "Plan РАЗОМ",
                  "Delta_РАЗОМ": "Δ РАЗОМ", "Pct_РАЗОМ": "% відхил."}
        for m in range(1, 13):
            ml = MONTH_LABELS[m]
            tt_lbl.update({f"fact_{ml}": f"{ml} Fact", f"plan_{ml}": f"{ml} Plan",
                           f"delta_{ml}": f"{ml} Δ"})

        header_row = [tt_lbl.get(c, c) for c in tt_export_cols]
        for ci, h in enumerate(header_row, 1):
            c = ws_tt.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill = hdr_fill("5b2d8e")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        ws_tt.row_dimensions[1].height = 28
        ws_tt.column_dimensions["A"].width = 22
        for ci in range(2, len(header_row) + 1):
            scw(ws_tt, ci, 10)

        df_tt_sorted = df_tt_agg[tt_export_cols].sort_values(
            "Delta_РАЗОМ" if "Delta_РАЗОМ" in df_tt_agg.columns else tt_export_cols[1],
            ascending=False
        ).reset_index(drop=True)

        for ri, row in df_tt_sorted.iterrows():
            for ci, col_name in enumerate(tt_export_cols, 1):
                v = row[col_name]
                c = ws_tt.cell(row=ri + 2, column=ci)
                c.border = thin_border()
                c.font   = Font(name="Arial", size=9)
                if col_name == "ТТ":
                    c.value = str(v) if pd.notna(v) else ""
                    c.alignment = Alignment(horizontal="left")
                elif col_name == "Pct_РАЗОМ":
                    c.value = float(v) / 100 if pd.notna(v) else None
                    c.number_format = PCT_FMT
                    c.alignment = Alignment(horizontal="right")
                    if pd.notna(v):
                        c.font = Font(name="Arial", size=9,
                                      color="C0392B" if v > 0 else ("2E7D32" if v < 0 else "000000"))
                else:
                    c.value = float(v) if pd.notna(v) else None
                    c.number_format = NUM_FMT
                    c.alignment = Alignment(horizontal="right")
                    if pd.notna(v) and isinstance(v, (int, float)) and v < 0:
                        c.font = Font(name="Arial", size=9, color="C0392B")

        total_ri = len(df_tt_sorted) + 2
        for ci, col_name in enumerate(tt_export_cols, 1):
            c = ws_tt.cell(row=total_ri, column=ci)
            c.border = thin_border()
            c.font   = Font(bold=True, name="Arial", size=9)
            c.fill   = hdr_fill("e8d5f5")
            if col_name == "ТТ":
                c.value = "🟰 РАЗОМ"
                c.alignment = Alignment(horizontal="left")
            elif col_name == "Pct_РАЗОМ":
                plan_s = df_tt_sorted["Plan_РАЗОМ"].sum() if "Plan_РАЗОМ" in df_tt_sorted else 0
                fact_s = df_tt_sorted["Fact_РАЗОМ"].sum() if "Fact_РАЗОМ" in df_tt_sorted else 0
                c.value = (fact_s / plan_s - 1) if plan_s != 0 else None
                c.number_format = PCT_FMT
                c.alignment = Alignment(horizontal="right")
            else:
                s = df_tt_sorted[col_name].sum() if col_name in df_tt_sorted else 0
                c.value = float(s) if pd.notna(s) else None
                c.number_format = NUM_FMT
                c.alignment = Alignment(horizontal="right")

    # ── 5. Heatmap ──────────────────────────────────────────────────────────
    if group_factors:
        heat, tt_table, val_col = build_heat_data(
            df, df_filtered, col_tt, col_article, col_month, col_value,
            col_plf, group_factors, articles_to_show, mode
        )
        ws_h = wb.create_sheet("Heatmap")
        ws_h.freeze_panes = "B2"
        heat_header = [col_tt] + MONTHS_LIST + ["РАЗОМ"]
        for ci, h in enumerate(heat_header, 1):
            c = ws_h.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill = hdr_fill("5b2d8e")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()

        try:
            import matplotlib.pyplot as plt
            import matplotlib.colors as mcolors
            flat = heat.values.flatten().astype(float)
            flat = flat[~np.isnan(flat)]
            norm = mcolors.Normalize(vmin=float(np.nanmin(flat)), vmax=float(np.nanmax(flat)))
            cmap_obj = plt.get_cmap("RdYlGn_r")
            use_cmap = True
        except Exception:
            use_cmap = False

        for ri, (idx, row) in enumerate(heat.iterrows(), 2):
            ws_h.cell(row=ri, column=1, value=idx).border = thin_border()
            ws_h.cell(row=ri, column=1).font = Font(name="Arial", size=9)
            for ci, col_name in enumerate(heat.columns, 2):
                v = row[col_name]
                c = ws_h.cell(row=ri, column=ci)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="right")
                if pd.isna(v):
                    c.fill = hdr_fill("FFFFFF")
                else:
                    c.value = float(v)
                    c.number_format = NUM_FMT
                    if use_cmap:
                        rgba = cmap_obj(norm(float(v)))
                        hx = "{:02X}{:02X}{:02X}".format(
                            int(rgba[0]*255), int(rgba[1]*255), int(rgba[2]*255))
                        c.fill = hdr_fill(hx)
                    c.font = Font(name="Arial", size=9, color="000000")

        ws_h.column_dimensions["A"].width = 20
        for ci in range(2, len(heat_header) + 1):
            scw(ws_h, ci, 11)

        ws_top    = wb.create_sheet("TOP_ANTITOP")
        sum_val   = tt_table.groupby(col_tt)[val_col].sum().reset_index()
        top_df    = sum_val.sort_values(val_col, ascending=True).head(50)
        anti_df   = sum_val.sort_values(val_col, ascending=False).head(50)

        def write_block(start_col, title_text, df_block, cmap_name):
            tc2 = ws_top.cell(row=1, column=start_col, value=title_text)
            tc2.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            tc2.fill = hdr_fill("5b2d8e")
            ws_top.merge_cells(start_row=1, start_column=start_col,
                               end_row=1, end_column=start_col + 1)
            for ci2, h2 in enumerate([col_tt, val_col], start_col):
                c = ws_top.cell(row=2, column=ci2, value=h2)
                c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                c.fill = hdr_fill("2e7d32")
                c.border = thin_border()
            try:
                import matplotlib.pyplot as plt
                import matplotlib.colors as mcolors
                vals_arr = df_block[val_col].values.astype(float)
                nm   = mcolors.Normalize(vmin=float(np.nanmin(vals_arr)),
                                         vmax=float(np.nanmax(vals_arr)))
                cm2  = plt.get_cmap(cmap_name)
                use_c = True
            except Exception:
                use_c = False
            for ri2, row2 in enumerate(df_block.itertuples(index=False), 3):
                tt_v  = getattr(row2, col_tt, "")
                val_v = getattr(row2, val_col, 0)
                ws_top.cell(row=ri2, column=start_col, value=tt_v).border = thin_border()
                ws_top.cell(row=ri2, column=start_col).font = Font(name="Arial", size=9)
                c2 = ws_top.cell(row=ri2, column=start_col + 1,
                                  value=float(val_v) if pd.notna(val_v) else None)
                c2.number_format = NUM_FMT
                c2.border = thin_border()
                c2.alignment = Alignment(horizontal="right")
                if use_c and pd.notna(val_v):
                    rgba2 = cm2(nm(float(val_v)))
                    hx2 = "{:02X}{:02X}{:02X}".format(
                        int(rgba2[0]*255), int(rgba2[1]*255), int(rgba2[2]*255))
                    c2.fill = hdr_fill(hx2)
                    lum = (0.299*int(rgba2[0]*255) + 0.587*int(rgba2[1]*255)
                           + 0.114*int(rgba2[2]*255)) / 255
                    c2.font = Font(name="Arial", size=9,
                                   color="000000" if lum > 0.5 else "FFFFFF")
            ws_top.column_dimensions[get_column_letter(start_col)].width = 22
            ws_top.column_dimensions[get_column_letter(start_col + 1)].width = 14

        write_block(1, "✅ TOP (економія)",      top_df,  "RdYlGn")
        write_block(4, "❌ ANTITOP (переліміт)", anti_df, "RdYlGn_r")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="СІМІ Dashboard", layout="wide")
    st.markdown("""
    <style>
    .simi-header { background:linear-gradient(90deg,#5b2d8e 0%,#7b52ae 100%);
                   color:white;padding:10px 18px;border-radius:4px;margin-bottom:4px; }
    .simi-logo   { font-size:2rem;font-weight:900;color:#f0c000;letter-spacing:1px;
                   margin-right:24px;vertical-align:middle; }
    .simi-store  { font-size:1.1rem;font-weight:700;color:white;vertical-align:middle; }
    .simi-meta-grid { display:grid;grid-template-columns:repeat(6,1fr);gap:4px 12px;margin-top:6px; }
    .simi-meta-item { font-size:0.78rem;color:#e0d0f8; }
    .simi-meta-val  { font-size:0.85rem;font-weight:700;color:white; }
    .article-selector { background:#f4f0fa;border:2px solid #5b2d8e;
                        border-radius:8px;padding:12px 16px;margin-bottom:14px; }
    .block-sep      { border-top:2px solid #5b2d8e;margin:16px 0 10px 0; }
    .block-sep-teal { border-top:2px solid #0d7377;margin:16px 0 10px 0; }
    .ratio-section-banner { background:linear-gradient(90deg,#085f63 0%,#0d7377 100%);
                            color:white;padding:8px 18px;border-radius:4px;
                            margin:8px 0 6px 0;font-size:0.95rem;font-weight:700; }
    div[data-testid="stButton"]>button[kind="primary"]   { background-color:#5b2d8e!important;color:white!important;border:2px solid #5b2d8e!important;font-weight:700!important;font-size:0.72rem!important; }
    div[data-testid="stButton"]>button[kind="secondary"] { background-color:#f4f0fa!important;color:#5b2d8e!important;border:1px solid #c9b6e8!important;font-size:0.72rem!important; }
    div[data-testid="stButton"]>button[kind="secondary"]:hover { background-color:#e8d5f5!important;border-color:#5b2d8e!important; }
    </style>""", unsafe_allow_html=True)

    file = st.file_uploader("📂 Завантажте Excel", type=["xlsx", "xlsb"])
    if file is None:
        st.info("Завантажте Excel-файл для початку роботи.")
        st.stop()

    file_bytes = file.read()
    xl         = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet_name = st.selectbox("Аркуш", xl.sheet_names)
    df         = load_excel(file_bytes, file.name, sheet_name)
    cols       = df.columns.tolist()

    with st.expander("⚙️ Налаштування колонок", expanded=True):
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            col_tt   = st.selectbox("TT (Магазин)", cols)
            col_year = st.selectbox("Year",          cols)
        with c2:
            col_month   = st.selectbox("Month",    cols)
            col_value   = st.selectbox("Значення", cols)
        with c3:
            col_plf     = st.selectbox("PL / F",         cols)
            col_article = st.selectbox("Стаття бюджету", cols)
        with c4:
            col_level0 = st.selectbox("Level_0", cols)
            ratio_candidates = ["— не обрано —"] + cols
            default_ratio_idx = 0
            for i, c in enumerate(cols):
                if "%" in c and "то" in c.lower():
                    default_ratio_idx = i + 1
                    break
            col_ratio = st.selectbox(
                "% в ТО без акцизу та без ПДВ", ratio_candidates,
                index=default_ratio_idx,
                help="Колонка з відсотком % в ТО. Оберіть 'не обрано' щоб приховати блок."
            )
            if col_ratio == "— не обрано —":
                col_ratio = None

    with st.expander("🏪 Колонки шапки магазину", expanded=False):
        sh1, sh2, sh3 = st.columns(3)
        with sh1:
            col_city  = st.selectbox("Місто",  ["—"] + cols)
            col_area  = st.selectbox("Площа",  ["—"] + cols)
        with sh2:
            col_format = st.selectbox("Формат ТО",   ["—"] + cols)
            col_mega   = st.selectbox("Мегасегмент", ["—"] + cols)
        with sh3:
            col_rik = st.selectbox("Рік",            ["—"] + cols)
            col_mis = st.selectbox("Місяць (шапка)", ["—"] + cols)

    # Sidebar
    st.sidebar.markdown("## 🔍 Фільтри")
    year_val   = st.sidebar.multiselect("Year",    sorted(df[col_year].dropna().unique(),   key=str))
    month_val  = st.sidebar.multiselect("Month",   sorted(df[col_month].dropna().unique(),  key=str))
    level0_val = st.sidebar.multiselect("Level_0", sorted(df[col_level0].dropna().unique(), key=str))

    st.sidebar.markdown("### ➕ Додаткові фільтри")
    fixed_cols  = {col_tt, col_year, col_month, col_level0}
    extra_filters = {}

    remaining = [c for c in cols if c not in fixed_cols]
    for i in range(1, 4):
        key_col = f"extra_filter_col{i}"
        key_val = f"extra_filter_val{i}"
        prev_extra = list(extra_filters.keys())
        options    = ["— не обрано —"] + [c for c in remaining if c not in prev_extra]
        chosen_col = st.sidebar.selectbox(f"Стовпець {i}", options, key=key_col)
        if chosen_col != "— не обрано —":
            chosen_val = st.sidebar.multiselect(
                f"Значення «{chosen_col}»",
                sorted(df[chosen_col].dropna().unique(), key=str), key=key_val,
            )
            if chosen_val:
                extra_filters[chosen_col] = chosen_val
        else:
            break

    # Pre-filter for TT list
    df_pre = df.copy()
    if year_val:   df_pre = df_pre[df_pre[col_year].isin(year_val)]
    if month_val:  df_pre = df_pre[df_pre[col_month].isin(month_val)]
    if level0_val: df_pre = df_pre[df_pre[col_level0].isin(level0_val)]
    for col_e, vals_e in extra_filters.items():
        df_pre = df_pre[df_pre[col_e].isin(vals_e)]

    visible_tts = sorted(df_pre[col_tt].dropna().unique(), key=str)

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🏪 ТТ за поточними фільтрами")
    if visible_tts:
        st.sidebar.caption(f"Знайдено: {len(visible_tts)} магазинів")
        tt_search    = st.sidebar.text_input("🔎 Пошук ТТ", value="",
                                              placeholder="Введіть назву...", key="tt_search")
        filtered_tts = [tt for tt in visible_tts if tt_search.lower() in str(tt).lower()] \
                       if tt_search else visible_tts
        b1, b2 = st.sidebar.columns(2)
        with b1:
            if st.button("✅ Всі", key="tt_select_all", use_container_width=True):
                st.session_state["tt_multiselect"] = filtered_tts
        with b2:
            if st.button("✖ Жодного", key="tt_clear_all", use_container_width=True):
                st.session_state["tt_multiselect"] = []
        tt_val = st.sidebar.multiselect(
            "Оберіть ТТ:", options=filtered_tts,
            default=st.session_state.get("tt_multiselect", []), key="tt_multiselect",
        )
    else:
        st.sidebar.warning("Немає ТТ за обраними фільтрами.")
        tt_val = []

    st.sidebar.markdown("---")
    mode       = st.sidebar.selectbox("Mode (Heatmap)",       ["Delta", "Delta %", "Z-score", "Fact", "Average"])
    ratio_mode = st.sidebar.selectbox("Mode (% в ТО Heatmap)", ["Delta", "Delta %", "Fact", "Average"], key="ratio_mode")
    
    options = [c for c in df.columns if c not in [col_value, col_plf, col_article]]

    group_factors = st.sidebar.multiselect(
    "Фактори групування (Average/Std)",
    options=options,
    default=[col_tt] if col_tt in options else [],   # перевірка, чи є col_tt серед options
    placeholder="Оберіть стовпці"
    )


    st.sidebar.markdown("---")
    st.sidebar.markdown("### 👁️ Відображення")
    show_ratio_section = st.sidebar.checkbox("Показати блок «% в ТО»", value=True,
                                              key="show_ratio_section") if col_ratio else False
    show_ratio_heatmap = st.sidebar.checkbox("Показати Heatmap % в ТО", value=True,
                                              key="show_ratio_heatmap") if col_ratio else False

    df[col_value] = pd.to_numeric(df[col_value], errors="coerce")

    def apply_filters(d):
        if tt_val:     d = d[d[col_tt].isin(tt_val)]
        if year_val:   d = d[d[col_year].isin(year_val)]
        if month_val:  d = d[d[col_month].isin(month_val)]
        if level0_val: d = d[d[col_level0].isin(level0_val)]
        for col_e, vals_e in extra_filters.items():
            d = d[d[col_e].isin(vals_e)]
        return d

    df_filtered = apply_filters(df).copy()

    def get_meta(col):
        if col == "—":
            return "—"
        vals = df_filtered[col].dropna().unique()
        return str(vals[0]) if len(vals) else "—"

    articles_all = sorted(df[col_article].dropna().unique(), key=str)

    st.markdown('<div class="article-selector">', unsafe_allow_html=True)
    sel_col1, sel_col2, sel_col3 = st.columns([3, 1, 1])
    with sel_col1:
        st.markdown("**🎯 Стаття бюджету для аналізу**")
        selected_article = st.selectbox("article_selector", articles_all,
                                         key="global_article", label_visibility="collapsed")
    with sel_col2:
        st.markdown("&nbsp;")
        show_all = st.checkbox("Показати всі статті", value=False, key="show_all")
    with sel_col3:
        st.markdown("&nbsp;")
        multi_sel = st.multiselect("Або обери кілька:", articles_all,
                                    default=[], key="multi_article")
    st.markdown('</div>', unsafe_allow_html=True)

    articles_to_show = (articles_all if show_all
                        else multi_sel if multi_sel
                        else [selected_article])

    if len(articles_to_show) == 1:
        st.info(f"📌 Показується стаття: **{articles_to_show[0]}**")
    else:
        st.info(f"📌 Показується {len(articles_to_show)} статей: {', '.join(articles_to_show)}")

    store_name = ", ".join(str(v) for v in tt_val) if tt_val else "Всі магазини"
    st.markdown(f"""
    <div class="simi-header">
      <span class="simi-logo">СіМі</span>
      <span class="simi-store">{store_name}</span>
      <div class="simi-meta-grid">
        <div><span class="simi-meta-item">Місто </span><span class="simi-meta-val">{get_meta(col_city)}</span></div>
        <div><span class="simi-meta-item">Площа </span><span class="simi-meta-val">{get_meta(col_area)}</span></div>
        <div><span class="simi-meta-item">Формат ТО </span><span class="simi-meta-val">{get_meta(col_format)}</span></div>
        <div><span class="simi-meta-item">Мегасегмент </span><span class="simi-meta-val">{get_meta(col_mega)}</span></div>
        <div><span class="simi-meta-item">Рік </span><span class="simi-meta-val">{get_meta(col_rik)}</span></div>
        <div><span class="simi-meta-item">Місяць </span><span class="simi-meta-val">{get_meta(col_mis)}</span></div>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── Article blocks ────────────────────────────────────────────────────────
    for art_idx, article in enumerate(articles_to_show):
        st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)

        tdf = build_article_monthly(
            df, df_filtered, col_tt, col_article, col_month, col_value,
            col_plf, article, tt_val, group_factors
        )
        render_article_block(
            title=article, table_df=tdf,
            df=df, df_filtered=df_filtered,
            col_tt=col_tt, col_article=col_article,
            col_month=col_month, col_value=col_value, col_plf=col_plf,
            group_factors=group_factors, tt_val=tt_val, article_idx=art_idx,
        )

        if col_ratio and show_ratio_section:
            rdf = build_ratio_monthly(
                df_filtered, col_tt, col_article, col_month, col_ratio, col_plf,
                article, tt_val, df_all=df, group_factors=group_factors
            )
            st.markdown('<div class="block-sep-teal"></div>', unsafe_allow_html=True)
            render_ratio_article_block(
                title=article, table_df=rdf,
                df=df, df_filtered=df_filtered,
                col_tt=col_tt, col_article=col_article,
                col_month=col_month, col_ratio=col_ratio, col_plf=col_plf,
                tt_val=tt_val, article_idx=art_idx, group_factors=group_factors,
            )

    # ── Pivot table ───────────────────────────────────────────────────────────
    st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)
    st.subheader("📋 Зведена таблиця")
    pivot_metric = st.radio("Метрика", ["Fact", "Plan", "Delta (Fact-Plan)"], horizontal=True)
    col_map_d    = {"Fact": "Fact", "Plan": "Plan", "Delta (Fact-Plan)": "Delta"}
    metric_col   = col_map_d[pivot_metric]

    rows_pivot = []
    for article in articles_to_show:
        tdf = build_article_monthly(df, df_filtered, col_tt, col_article,
                                    col_month, col_value, col_plf, article, tt_val, group_factors)
        row = {"Стаття": article}
        for m in range(1, 13):
            row[MONTH_LABELS[m]] = tdf.loc[m, metric_col]
        row["РАЗОМ"] = sum(tdf.loc[m, metric_col] for m in range(1, 13))
        rows_pivot.append(row)

    pivot_df = pd.DataFrame(rows_pivot).set_index("Стаття")
    cmap_p   = "RdYlGn_r" if pivot_metric == "Delta (Fact-Plan)" else "Blues"
    st.dataframe(
        pivot_df.style
            .background_gradient(cmap=cmap_p, axis=None)
            .format(lambda v: f"{v:,.0f}".replace(",", " ") if pd.notna(v) else "-", na_rep="-"),
        use_container_width=True,
    )

    # ── % в ТО зведена ───────────────────────────────────────────────────────
    if col_ratio and show_ratio_section:
        st.markdown('<div class="block-sep-teal"></div>', unsafe_allow_html=True)
        st.markdown('<div class="ratio-section-banner">📊 Зведена таблиця — % в ТО без акцизу та без ПДВ</div>',
                    unsafe_allow_html=True)
        ratio_pivot_metric = st.radio("Метрика (% в ТО)", ["Fact", "Plan", "Average", "Delta"],
                                       horizontal=True, key="ratio_pivot_metric")
        rows_ratio_pivot = []
        for article in articles_to_show:
            rdf = build_ratio_monthly(
                df_filtered, col_tt, col_article, col_month, col_ratio, col_plf,
                article, tt_val, df_all=df, group_factors=group_factors
            )
            row  = {"Стаття": article}
            vals = [rdf.loc[m, ratio_pivot_metric] for m in range(1, 13)]
            for m in range(1, 13):
                row[MONTH_LABELS[m]] = rdf.loc[m, ratio_pivot_metric]
            nz = [v for v in vals if v != 0]
            row["Серед."] = np.mean(nz) if nz else 0.0
            rows_ratio_pivot.append(row)

        ratio_pivot_df = pd.DataFrame(rows_ratio_pivot).set_index("Стаття")
        st.dataframe(
            ratio_pivot_df.style
                .background_gradient(cmap="RdYlGn_r", axis=None)
                .format(lambda v: f"{v:.2f}%" if pd.notna(v) else "-", na_rep="-"),
            use_container_width=True,
        )

    # ── TT Pivot ──────────────────────────────────────────────────────────────
    st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)
    st.subheader("📋 Зведена таблиця в розрізі ТТ")
    tt_pivot_metric = st.radio("Метрика (ТТ)", ["Fact", "Plan", "Delta (Fact-Plan)"],
                                horizontal=True, key="tt_pivot_metric")
    tt_metric_col   = col_map_d[tt_pivot_metric]
    show_pct        = st.checkbox("Показати % відхилення (Fact vs Plan)", value=True, key="show_pct")
    show_months     = st.checkbox("Розгорнути по місяцях", value=False, key="tt_show_months")

    df_tt_agg = build_tt_pivot(
        df_filtered, col_tt, col_article, col_month, col_value, col_plf, articles_to_show
    )

    if df_tt_agg.empty:
        st.info("Немає даних для побудови таблиці по ТТ.")
    else:
        if show_months:
            display_cols = ["ТТ"]
            col_labels   = {"ТТ": "ТТ"}
            for m in range(1, 13):
                ml = MONTH_LABELS[m]
                if tt_metric_col in ("Fact", "Delta"):
                    display_cols.append(f"fact_{ml}")
                    col_labels[f"fact_{ml}"] = f"{ml} Fact"
                if tt_metric_col == "Plan":
                    display_cols.append(f"plan_{ml}")
                    col_labels[f"plan_{ml}"] = f"{ml} Plan"
                if show_pct and tt_metric_col != "Plan":
                    display_cols.append(f"pct_{ml}")
                    col_labels[f"pct_{ml}"] = f"{ml} %"
        else:
            display_cols = ["ТТ"]
            col_labels   = {"ТТ": "ТТ"}

        if tt_metric_col == "Fact":
            display_cols += ["Fact_РАЗОМ"]
            col_labels["Fact_РАЗОМ"] = "Fact РАЗОМ"
        elif tt_metric_col == "Plan":
            display_cols += ["Plan_РАЗОМ"]
            col_labels["Plan_РАЗОМ"] = "Plan РАЗОМ"
        else:
            display_cols += ["Fact_РАЗОМ", "Plan_РАЗОМ", "Delta_РАЗОМ"]
            col_labels.update({"Fact_РАЗОМ": "Fact РАЗОМ",
                               "Plan_РАЗОМ": "Plan РАЗОМ", "Delta_РАЗОМ": "Δ РАЗОМ"})

        if show_pct:
            display_cols.append("Pct_РАЗОМ")
            col_labels["Pct_РАЗОМ"] = "% відхил."

        df_display = df_tt_agg[display_cols].rename(columns=col_labels).set_index("ТТ")

        sort_col_label = ("% відхил." if show_pct else
                          "Δ РАЗОМ"    if tt_metric_col == "Delta (Fact-Plan)" else
                          "Fact РАЗОМ" if tt_metric_col == "Fact" else "Plan РАЗОМ")
        if sort_col_label in df_display.columns:
            df_display = df_display.sort_values(sort_col_label, ascending=True)

        total_row = df_display.sum(numeric_only=True)
        if "% відхил." in df_display.columns:
            plan_sum = df_tt_agg["Plan_РАЗОМ"].sum()
            fact_sum = df_tt_agg["Fact_РАЗОМ"].sum()
            total_row["% відхил."] = (fact_sum / plan_sum - 1) * 100 if plan_sum != 0 else None
        total_row.name = "🟰 РАЗОМ"
        df_display     = pd.concat([df_display, total_row.to_frame().T])

        pct_cols = [c for c in df_display.columns if "%" in c]
        num_cols = [c for c in df_display.columns if "%" not in c]
        fmt_dict = {c: (lambda v: f"{v:,.0f}".replace(",", " ") if pd.notna(v) else "-") for c in num_cols}
        fmt_dict.update({c: (lambda v: f"{'+' if v > 0 else ''}{v:.1f}%" if pd.notna(v) else "-")
                         for c in pct_cols})

        styled = df_display.style.format(fmt_dict, na_rep="-")
        if pct_cols:
            styled = styled.background_gradient(cmap="RdYlGn_r",
                        subset=pd.IndexSlice[df_display.index[:-1], pct_cols], axis=None)
        delta_cols = [c for c in num_cols if "Δ" in c]
        other_cols = [c for c in num_cols if "Δ" not in c]
        if delta_cols:
            styled = styled.background_gradient(cmap="RdYlGn_r",
                        subset=pd.IndexSlice[df_display.index[:-1], delta_cols], axis=None)
        if other_cols:
            styled = styled.background_gradient(cmap="Blues",
                        subset=pd.IndexSlice[df_display.index[:-1], other_cols], axis=None)
        styled = styled.apply(
            lambda row: ["font-weight:bold;border-top:2px solid #5b2d8e;" for _ in row]
            if row.name == "🟰 РАЗОМ" else ["" for _ in row], axis=1
        )
        st.dataframe(styled, use_container_width=True, height=500)
        st.download_button(
            "⬇️ Завантажити CSV (ТТ-зведена)",
            data=df_display.to_csv(encoding="utf-8-sig").encode("utf-8-sig"),
            file_name="tt_pivot.csv", mime="text/csv", key="tt_pivot_csv",
        )

    # ── Heatmap ───────────────────────────────────────────────────────────────
    st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)
    st.subheader("🌡️ Карта аномалій по магазинах")

    if group_factors:
        heat, tt_table, val_col = build_heat_data(
            df, df_filtered, col_tt, col_article, col_month, col_value,
            col_plf, group_factors, articles_to_show, mode
        )
        st.dataframe(
            heat.style
                .background_gradient(cmap="RdYlGn_r", axis=None)
                .highlight_null(color="white")
                .format(lambda v: f"{v:,.0f}".replace(",", " ") if pd.notna(v) else "", na_rep=""),
            use_container_width=True,
        )

        st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)
        st.subheader("🏆 TOP / ANTITOP магазинів")
        sum_val = tt_table.groupby(col_tt)[val_col].sum().reset_index()
        n_tt    = st.slider("Кількість магазинів", 1, 100, 10)
        top     = sum_val.sort_values(val_col, ascending=True).head(n_tt)
        antitop = sum_val.sort_values(val_col, ascending=False).head(n_tt)
        fmt_abs = lambda v: f"{v:,.0f}".replace(",", " ") if pd.notna(v) else "-"
        ca, cb  = st.columns(2)
        with ca:
            st.write("✅ Top (економія)")
            st.dataframe(top.style.background_gradient(cmap="RdYlGn", subset=[val_col])
                            .format({val_col: fmt_abs}))
        with cb:
            st.write("❌ Antitop (переліміт)")
            st.dataframe(antitop.style.background_gradient(cmap="RdYlGn_r", subset=[val_col])
                                 .format({val_col: fmt_abs}))
    else:
        st.info("Оберіть фактори групування в боковому меню для побудови Heatmap.")

    # ── % в ТО Heatmap ────────────────────────────────────────────────────────
    if col_ratio and show_ratio_heatmap:
        st.markdown('<div class="block-sep-teal"></div>', unsafe_allow_html=True)
        render_ratio_heatmap_section(
            df, df_filtered, col_tt, col_article, col_month,
            col_ratio, col_plf, articles_to_show, ratio_mode,
            group_factors=group_factors,
        )

    # ── Export ────────────────────────────────────────────────────────────────
    st.markdown('<div class="block-sep"></div>', unsafe_allow_html=True)
    st.subheader("📥 Експорт в Excel")

    export_sections = ["✅ Зведена таблиця (статті)", "✅ Листи по кожній статті (з графіком)"]
    if df_tt_agg is not None and not df_tt_agg.empty:
        export_sections.append("✅ ТТ-Зведена таблиця")
    if col_ratio:
        export_sections.append("✅ % в ТО — зведена таблиця")
    if group_factors:
        export_sections += ["✅ Heatmap аномалій", "✅ TOP / ANTITOP магазинів"]

    st.markdown("**Файл міститиме аркуші:**")
    for s in export_sections:
        st.markdown(f"- {s}")

    st.download_button(
        label="⬇️ Скачати дашборд як Excel",
        data=export_excel(
            df, df_filtered, col_tt, col_article, col_month, col_value,
            col_plf, articles_to_show, tt_val, group_factors, metric_col,
            mode, pivot_df, df_tt_agg=df_tt_agg, col_ratio=col_ratio,
        ),
        file_name="simi_dashboard.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
